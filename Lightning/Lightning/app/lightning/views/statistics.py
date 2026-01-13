from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Exists, OuterRef

from app.lightning.models.lightning_models import Lightning
from app.lightning.models.lightning_data_models import LightningRead, LightningTest
from app.lightning.models.lightning_test_models import LightningQuestion
from app.bot.models import TelegramUser
from app.organization.models.company_models import Department, JobTitle


@staff_member_required
def lightning_statistics_view(request):
    # Молнии для выбора (минимально необходимые поля)
    lightnings = (
        Lightning.objects.filter(is_draft=False)
        .only("id", "name")
        .order_by("-created_at")
    )

    # Выбранная молния
    selected_lightning_id = request.GET.get("lightning_id")
    selected_lightning = None
    if selected_lightning_id and selected_lightning_id.isdigit():
        selected_lightning = lightnings.filter(id=int(selected_lightning_id)).first()

    # Блок статистики
    if selected_lightning:
        total_participants = LightningRead.objects.filter(
            lightning=selected_lightning, user__state=TelegramUser.STATE_ACTIVE
        ).count()
        read_count = LightningRead.objects.filter(
            lightning=selected_lightning,
            is_read=True,
            user__state=TelegramUser.STATE_ACTIVE,
        ).count()
        not_read_count = LightningRead.objects.filter(
            lightning=selected_lightning,
            is_read=False,
            user__state=TelegramUser.STATE_ACTIVE,
        ).count()
        completed_tests = LightningTest.objects.filter(
            lightning=selected_lightning,
            complete=True,
            user__state=TelegramUser.STATE_ACTIVE,
        ).count()
        not_completed_tests = LightningTest.objects.filter(
            lightning=selected_lightning,
            complete=False,
            user__state=TelegramUser.STATE_ACTIVE,
        ).count()
    else:
        total_participants = 0
        read_count = 0
        not_read_count = 0
        completed_tests = 0
        not_completed_tests = 0

    # Таблица пользователей (с пагинацией)
    user_list = []
    page_obj = None
    user_filter = request.GET.get("user_filter")
    selected_department_ids = request.GET.getlist("department_id")
    selected_job_title_ids = request.GET.getlist("job_title_id")
    selected_location_ids = request.GET.getlist("location_id")
    if user_filter and selected_lightning:
        if user_filter == "all":
            user_ids = LightningRead.objects.filter(
                lightning=selected_lightning
            ).values_list("user_id", flat=True)
        elif user_filter == "read":
            user_ids = LightningRead.objects.filter(
                lightning=selected_lightning, is_read=True
            ).values_list("user_id", flat=True)
        elif user_filter == "not_read":
            user_ids = LightningRead.objects.filter(
                lightning=selected_lightning, is_read=False
            ).values_list("user_id", flat=True)
        elif user_filter == "completed_tests":
            user_ids = LightningTest.objects.filter(
                lightning=selected_lightning, complete=True
            ).values_list("user_id", flat=True)
        elif user_filter == "not_completed_tests":
            user_ids = LightningTest.objects.filter(
                lightning=selected_lightning, complete=False
            ).values_list("user_id", flat=True)
        else:
            user_ids = []

        users_qs = (
            TelegramUser.objects.filter(
                id__in=user_ids, state=TelegramUser.STATE_ACTIVE
            )
            .select_related("department", "job_title")
            .order_by("full_name")
        )
        paginator = Paginator(users_qs, 25)
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)
        user_list = page_obj.object_list

    return render(
        request,
        "lightning_statistics.html",
        {
            "lightnings": lightnings,
            "selected_lightning": selected_lightning,
            "total_participants": total_participants,
            "read_count": read_count,
            "not_read_count": not_read_count,
            "completed_tests": completed_tests,
            "not_completed_tests": not_completed_tests,
            "user_list": user_list,
            "page_obj": page_obj,
            "selected_department_ids": selected_department_ids,
            "selected_job_title_ids": selected_job_title_ids,
            "selected_location_ids": selected_location_ids,
        },
    )


@staff_member_required
def lightnings_vs_users_statistics_view(request):
    """
    Отчёт «Молнии × Пользователи»

    Логика фильтров:
    - Подразделение (multiple, только родительские): если выбрано ХОТЬ ОДНО подразделение —
      включаем КАЖДОГО выбранного родителя и ВСЕХ его потомков. Выбор Локаций в этом случае ИГНОРИРУЕТСЯ.
    - Локация (multiple, только листья): применяется ТОЛЬКО если подразделения не выбраны.
    - Должность (multiple, OR).
    - Экспорт XLSX без пагинации; на странице считаем статусы только по видимым пользователям.
    """

    # ---------- столбцы (молнии) ----------
    all_lightnings = (
        Lightning.objects.filter(is_draft=False)
        .only("id", "name", "created_at")
        .order_by("-created_at")
    )

    # ---------- фильтры из запроса ----------
    lightning_id = request.GET.get("lightning_id")
    department_ids_raw = request.GET.getlist("department_id")  # multiple (родители)
    job_title_ids_raw = request.GET.getlist("job_title_id")  # multiple
    location_ids_raw = request.GET.getlist("location_id")  # multiple (листья)
    user_query = request.GET.get("user_query")
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")

    def parse_date(value: str | None):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except Exception:
            return None

    def to_ints(values):
        out = []
        for v in values:
            v = (v or "").strip()
            if v.isdigit():
                out.append(int(v))
        return out

    date_from = parse_date(date_from_str)
    date_to = parse_date(date_to_str)

    has_active_filters = any(
        [
            lightning_id,
            user_query,
            date_from_str,
            date_to_str,
            department_ids_raw,
            job_title_ids_raw,
            location_ids_raw,
        ]
    )

    # Применяем фильтры к молниям
    lightnings = all_lightnings
    if lightning_id and lightning_id.isdigit():
        lightnings = lightnings.filter(id=int(lightning_id))
    if date_from:
        lightnings = lightnings.filter(created_at__date__gte=date_from)
    if date_to:
        lightnings = lightnings.filter(created_at__date__lte=date_to)

    # ---------- пользователи (строки) ----------
    users_qs = (
        TelegramUser.objects.filter(state=TelegramUser.STATE_ACTIVE)
        .only("id", "full_name", "department", "job_title")
        .select_related("department", "department__parent", "job_title")
        .order_by("full_name")
    )

    # ========== МНОЖЕСТВЕННЫЕ ФИЛЬТРЫ ==========
    # 1) Подразделения-родители: если выбраны — включаем РОДИТЕЛЕЙ и ВСЕХ их потомков.
    dept_roots = set(to_ints(department_ids_raw))
    all_dept_ids = set()
    if dept_roots:
        all_dept_ids |= dept_roots  # включаем самих родителей
        to_check = list(dept_roots)
        while to_check:
            child_ids = set(
                Department.objects.filter(parent_id__in=to_check).values_list(
                    "id", flat=True
                )
            )
            child_ids -= all_dept_ids
            if not child_ids:
                break
            all_dept_ids |= child_ids
            to_check = list(child_ids)

        # Применяем фильтр по отделам (родители + их потомки)
        users_qs = users_qs.filter(department_id__in=all_dept_ids)

    else:
        # 2) Локации (только если подразделения не выбраны): валидируем, что это листья
        loc_ids_in = set(to_ints(location_ids_raw))
        if loc_ids_in:
            valid_location_ids = set(
                Department.objects.annotate(
                    has_children=Exists(
                        Department.objects.filter(parent_id=OuterRef("pk"))
                    )
                )
                .filter(has_children=False, id__in=loc_ids_in)
                .values_list("id", flat=True)
            )
            if valid_location_ids:
                users_qs = users_qs.filter(department_id__in=valid_location_ids)

    # 3) Должности — множественный OR
    job_title_ids = to_ints(job_title_ids_raw)
    if job_title_ids:
        users_qs = users_qs.filter(job_title_id__in=job_title_ids)

    # 4) Поиск по ФИО
    if user_query:
        users_qs = users_qs.filter(full_name__icontains=user_query.strip())

    # ---------- списки для выпадашек ----------
    # Локации (только ЛИСТЫ)
    locations = (
        Department.objects.annotate(
            has_children=Exists(Department.objects.filter(parent_id=OuterRef("pk")))
        )
        .filter(has_children=False)
        .only("id", "name")
        .order_by("name")
    )
    # Подразделения (только РОДИТЕЛИ)
    departments_parents = (
        Department.objects.annotate(
            has_children=Exists(Department.objects.filter(parent_id=OuterRef("pk")))
        )
        .filter(has_children=True)
        .only("id", "name")
        .order_by("name")
    )

    # ---------- базовая строка запроса без page/export ----------
    query_params = request.GET.copy()
    for k in ("page", "export"):
        if k in query_params:
            query_params.pop(k)
    base_query = query_params.urlencode()

    # ---------- Экспорт ----------
    export_fmt = request.GET.get("export")  # 'xlsx'
    if export_fmt:
        lightning_ids = list(lightnings.values_list("id", flat=True))
        user_ids = list(users_qs.values_list("id", flat=True))

        lightnings_with_questions = set(
            LightningQuestion.objects.filter(
                lightning_id__in=lightning_ids
            ).values_list("lightning_id", flat=True)
        )
        read_pairs = {
            f"{uid}_{lid}"
            for uid, lid in LightningRead.objects.filter(
                user_id__in=user_ids, lightning_id__in=lightning_ids, is_read=True
            ).values_list("user_id", "lightning_id")
        }
        test_complete_pairs = {
            f"{uid}_{lid}"
            for uid, lid in LightningTest.objects.filter(
                user_id__in=user_ids, lightning_id__in=lightning_ids, complete=True
            ).values_list("user_id", "lightning_id")
        }
        has_read_record_pairs = {
            f"{uid}_{lid}"
            for uid, lid in LightningRead.objects.filter(
                user_id__in=user_ids, lightning_id__in=lightning_ids
            ).values_list("user_id", "lightning_id")
        }

        use_xlsx = False
        if export_fmt == "xlsx":
            try:
                from openpyxl import Workbook  # noqa

                use_xlsx = True
            except Exception:
                use_xlsx = False

        # XLSX
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"
        ws.append(
            [
                "Сотрудник",
                "Должность",
                "Родитель",
                "Отдел",
                *list(lightnings.values_list("name", flat=True)),
            ]
        )

        for u in users_qs.iterator(chunk_size=2000):
            u_id = u.id
            dept = u.department
            jt = u.job_title
            row = [
                u.full_name,
                jt.name if jt else "",
                dept.parent.name if (dept and dept.parent) else "",
                dept.name if dept else "",
            ]
            for l_id in lightning_ids:
                key = f"{u_id}_{l_id}"
                if key not in has_read_record_pairs:
                    row.append("не получал")
                elif l_id not in lightnings_with_questions:
                    row.append("прочитал" if key in read_pairs else "не прочитал")
                else:
                    if key in test_complete_pairs:
                        row.append("сдал")
                    elif key in read_pairs:
                        row.append("прочитал")
                    else:
                        row.append("не прочитал")
            ws.append(row)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="lightnings_vs_users.xlsx"'
        )
        wb.save(response)
        return response

    # ---------- Обычный режим (рендер с пагинацией) ----------
    paginator = Paginator(users_qs, 25)
    page_number = request.GET.get("page")
    users_page = paginator.get_page(page_number)
    users = list(users_page.object_list)

    user_ids_page = [u.id for u in users]
    lightning_ids_page = list(lightnings.values_list("id", flat=True))

    lightnings_with_questions_page = list(
        LightningQuestion.objects.filter(
            lightning_id__in=lightning_ids_page
        ).values_list("lightning_id", flat=True)
    )
    read_pairs = {
        f"{uid}_{lid}"
        for uid, lid in LightningRead.objects.filter(
            user_id__in=user_ids_page, lightning_id__in=lightning_ids_page, is_read=True
        ).values_list("user_id", "lightning_id")
    }
    test_complete_pairs = {
        f"{uid}_{lid}"
        for uid, lid in LightningTest.objects.filter(
            user_id__in=user_ids_page,
            lightning_id__in=lightning_ids_page,
            complete=True,
        ).values_list("user_id", "lightning_id")
    }
    has_read_record_pairs = {
        f"{uid}_{lid}"
        for uid, lid in LightningRead.objects.filter(
            user_id__in=user_ids_page,
            lightning_id__in=lightning_ids_page,
        ).values_list("user_id", "lightning_id")
    }

    # Выбранные значения для multiple-селектов (для шаблона)
    selected_department_ids = request.GET.getlist("department_id")
    selected_job_title_ids = request.GET.getlist("job_title_id")
    selected_location_ids = request.GET.getlist("location_id")

    return render(
        request,
        "lightnings_vs_users_statistics.html",
        {
            "lightnings": lightnings,
            "filter_lightnings": all_lightnings,
            "users_page": users_page,
            "users": users,
            "read_pairs": read_pairs,
            "test_complete_pairs": test_complete_pairs,
            "departments": departments_parents,  # только родительские отделы
            "job_titles": JobTitle.objects.all().only("id", "name").order_by("name"),
            "locations": locations,  # только листья
            "base_query": base_query,
            "lightnings_with_questions": lightnings_with_questions_page,
            "has_active_filters": has_active_filters,
            "selected_department_ids": selected_department_ids,
            "selected_job_title_ids": selected_job_title_ids,
            "selected_location_ids": selected_location_ids,
            "has_read_record_pairs": has_read_record_pairs,
        },
    )


@staff_member_required
def lightnings_summary_by_departments_view(request):
    """
    Сводный отчёт: Молнии (строки) x Отделы (столбцы).
    Ячейка: количество "не ознакомленных".
    Критерий "не ознакомлен":
      - (is_read=False)
      - ИЛИ (is_read=True И (тест есть И complete=False))
      - (если теста нет или он сдан — считаем ознакомленным)
    """

    def to_ints(values):
        out = []
        for v in values:
            v = (v or "").strip()
            if v.isdigit():
                out.append(int(v))
        return out

    # 1. Молнии (фильтруем draft=False)
    qs = Lightning.objects.filter(is_draft=False).order_by("-created_at")

    # Фильтры
    lightning_ids_raw = request.GET.getlist("lightning_id")
    department_ids_raw = request.GET.getlist("department_id")  # multiple (родители)
    location_ids_raw = request.GET.getlist("location_id")  # multiple (листья)
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")

    def parse_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except Exception:
            return None

    date_from = parse_date(date_from_str)
    date_to = parse_date(date_to_str)

    lightning_ids_in = to_ints(lightning_ids_raw)
    if lightning_ids_in:
        qs = qs.filter(id__in=lightning_ids_in)

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    lightnings = list(qs)
    lightning_ids = [l.id for l in lightnings]

    # Для фильтра в шаблоне (список всех не-черновиков)
    all_lightnings = (
        Lightning.objects.filter(is_draft=False)
        .only("id", "name")
        .order_by("-created_at")
    )

    has_active_filters = any(
        [
            lightning_ids_raw,
            date_from_str,
            date_to_str,
            department_ids_raw,
            location_ids_raw,
        ]
    )

    # 2. Данные по статусам
    # Нам нужно знать для каждой (lightning_id, user_id):
    #   - есть ли запись в LightningRead (если нет - не отправлялась, не считаем)
    #   - is_read
    #   - complete (из LightningTest)
    #   - department_id юзера

    # Сначала возьмём всех LightningRead для выбранных молний + активных юзеров
    reads = LightningRead.objects.filter(
        lightning_id__in=lightning_ids, user__state=TelegramUser.STATE_ACTIVE
    )

    # ========== МНОЖЕСТВЕННЫЕ ФИЛЬТРЫ ПО ОТДЕЛАМ (User Filter) ==========
    # 1) Подразделения-родители: если выбраны — включаем РОДИТЕЛЕЙ и ВСЕХ их потомков.
    dept_roots = set(to_ints(department_ids_raw))
    all_dept_ids = set()
    if dept_roots:
        all_dept_ids |= dept_roots  # включаем самих родителей
        to_check = list(dept_roots)
        while to_check:
            child_ids = set(
                Department.objects.filter(parent_id__in=to_check).values_list(
                    "id", flat=True
                )
            )
            child_ids -= all_dept_ids
            if not child_ids:
                break
            all_dept_ids |= child_ids
            to_check = list(child_ids)

        reads = reads.filter(user__department_id__in=all_dept_ids)

    else:
        # 2) Локации (только если подразделения не выбраны): валидируем, что это листья
        loc_ids_in = set(to_ints(location_ids_raw))
        if loc_ids_in:
            valid_location_ids = set(
                Department.objects.annotate(
                    has_children=Exists(
                        Department.objects.filter(parent_id=OuterRef("pk"))
                    )
                )
                .filter(has_children=False, id__in=loc_ids_in)
                .values_list("id", flat=True)
            )
            if valid_location_ids:
                reads = reads.filter(user__department_id__in=valid_location_ids)

    # Выбираем нужные поля
    reads_data = reads.values(
        "lightning_id", "user_id", "is_read", "user__department_id"
    )

    # Собираем словарь: (lightning_id, user_id) -> {is_read, dept_id}
    # И сразу список пар (lightning_id, user_id), для которых is_read=True, чтобы проверить тесты
    user_lightning_map = {}
    check_test_keys = []

    # Также соберем все dept_id, которые участвуют в "не ознакомлен"
    # Но сначала надо вычислить статус "не ознакомлен".

    for r in reads_data:
        key = (r["lightning_id"], r["user_id"])
        user_lightning_map[key] = {
            "is_read": r["is_read"],
            "dept_id": r["user__department_id"],
            "not_read": False,  # пока false, вычислим ниже
        }
        if r["is_read"]:
            check_test_keys.append(key)
        else:
            # Если не прочитал - сразу должник
            user_lightning_map[key]["not_read"] = True

    # Теперь подгрузим тесты для тех, кто прочитал
    if check_test_keys:
        l_ids_test = set(k[0] for k in check_test_keys)
        u_ids_test = set(k[1] for k in check_test_keys)

        tests = LightningTest.objects.filter(
            lightning_id__in=l_ids_test, user_id__in=u_ids_test
        ).values("lightning_id", "user_id", "complete")

        test_map = {(t["lightning_id"], t["user_id"]): t["complete"] for t in tests}

        l_with_questions = set(
            LightningQuestion.objects.filter(lightning_id__in=l_ids_test).values_list(
                "lightning_id", flat=True
            )
        )

        for key in check_test_keys:
            lid, uid = key
            if lid in l_with_questions:
                # Тест нужен.
                # Должник ТОЛЬКО если запись ЕСТЬ и complete=False.
                is_complete = test_map.get(key)  # True/False/None
                if is_complete is False:
                    user_lightning_map[key]["not_read"] = True
            else:
                # Теста нет -> ознакомлен
                pass

    # 3. Агрегация данных
    involved_dept_ids = set()

    # Справочник для строк
    rows_map = {l.id: {"lightning": l, "total": 0, "by_dept": {}} for l in lightnings}

    for key, info in user_lightning_map.items():
        if info["not_read"]:
            lid, uid = key
            if lid not in rows_map:
                continue  # вдруг отфильтровано

            dept_id = info["dept_id"]
            # Если dept_id is None -> "-"
            d_key = dept_id if dept_id is not None else -1

            rows_map[lid]["total"] += 1
            rows_map[lid]["by_dept"][d_key] = rows_map[lid]["by_dept"].get(d_key, 0) + 1

            involved_dept_ids.add(d_key)

    # 4. Формируем колонки
    real_dept_ids = [d for d in involved_dept_ids if d != -1]
    depts_qs = Department.objects.filter(id__in=real_dept_ids).select_related("parent")

    dept_headers = []
    for d in depts_qs:
        dept_headers.append(
            {"id": d.id, "name": d.name, "full_str": f"{d.name} (ID: {d.id})"}
        )

    # Сортировка колонок (по имени)
    dept_headers.sort(key=lambda x: x["name"])

    # Если есть "без отдела", добавим в конец
    has_no_dept = -1 in involved_dept_ids
    if has_no_dept:
        dept_headers.append({"id": -1, "name": "-", "full_str": "-"})

    # 5. Итоговая строка и подготовка данных для шаблона
    total_row_data = {"total": 0, "by_dept": {h["id"]: 0 for h in dept_headers}}

    processed_rows = []
    for l in lightnings:
        row_data = rows_map[l.id]

        total_row_data["total"] += row_data["total"]
        for h in dept_headers:
            did = h["id"]
            val = row_data["by_dept"].get(did, 0)
            total_row_data["by_dept"][did] += val

        processed_rows.append(row_data)

    # Преобразуем словари by_dept в списки, соответствующие dept_headers
    final_rows = []
    for r in processed_rows:
        values = []
        for h in dept_headers:
            values.append(r["by_dept"].get(h["id"], 0))

        final_rows.append(
            {"lightning": r["lightning"], "total": r["total"], "values": values}
        )

    total_values = []
    for h in dept_headers:
        total_values.append(total_row_data["by_dept"].get(h["id"], 0))

    final_total_row = {"total": total_row_data["total"], "values": total_values}

    # ---------- Экспорт XLSX ----------
    if request.GET.get("export") == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный по отделам"

        # Заголовки
        headers = ["Молния", "Дата", "Всего должников"] + [
            h["full_str"] for h in dept_headers
        ]
        ws.append(headers)

        # Стили для заголовка
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="F2F2F2")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Строки данных
        for row in final_rows:
            l = row["lightning"]
            data = [
                l.name,
                l.created_at.strftime("%d.%m.%Y") if l.created_at else "",
                row["total"],
            ]
            # Значения по отделам (0 меняем на пустую строку для чистоты или оставляем 0)
            for val in row["values"]:
                data.append(val if val > 0 else "")

            ws.append(data)

        # Итого
        total_data = ["ИТОГО", "", final_total_row["total"]]
        for val in final_total_row["values"]:
            total_data.append(val if val > 0 else "")

        ws.append(total_data)

        # Жирный шрифт для итоговой строки
        last_row_idx = len(final_rows) + 2
        for cell in ws[last_row_idx]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E9ECEF")

        # Автоширина колонок
        for col_idx, cell in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="lightnings_summary_by_departments.xlsx"'
        )
        wb.save(response)
        return response

    # ---------- Данные для фильтров (аналогично users_vs_lightnings) ----------
    locations = (
        Department.objects.annotate(
            has_children=Exists(Department.objects.filter(parent_id=OuterRef("pk")))
        )
        .filter(has_children=False)
        .only("id", "name")
        .order_by("name")
    )
    departments_parents = (
        Department.objects.annotate(
            has_children=Exists(Department.objects.filter(parent_id=OuterRef("pk")))
        )
        .filter(has_children=True)
        .only("id", "name")
        .order_by("name")
    )

    # Выбранные ID для шаблона
    selected_lightning_ids = [str(i) for i in lightning_ids_in]
    selected_department_ids = department_ids_raw
    selected_location_ids = location_ids_raw

    # базовая строка запроса (для сброса пагинации, если бы она была, или для других ссылок)
    query_params = request.GET.copy()
    if "export" in query_params:
        query_params.pop("export")
    base_query = query_params.urlencode()

    return render(
        request,
        "lightnings_summary_by_departments.html",
        {
            "rows": final_rows,
            "dept_headers": dept_headers,
            "total_row": final_total_row,
            "filter_lightnings": all_lightnings,
            "has_active_filters": has_active_filters,
            # Новые данные для фильтров
            "departments": departments_parents,
            "locations": locations,
            "selected_lightning_ids": selected_lightning_ids,
            "selected_department_ids": selected_department_ids,
            "selected_location_ids": selected_location_ids,
            "base_query": base_query,
        },
    )
